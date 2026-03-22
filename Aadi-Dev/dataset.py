# ============================================================
# MEMBER 1 — Features F01 to F12
# Group A: Thermodynamic / Structural  (F01–F06)
# Group B: Electronic Structure        (F07–F12)
#
# INPUT  : UHTM_base_200.xlsx
# OUTPUT : UHTM_member1_F01_F12.xlsx
# ============================================================

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

np.random.seed(42)

# ── Load base file ─────────────────────────────────────────────────────────────
df = pd.read_excel("UHTM_base_200.xlsx", sheet_name="Base Data")

mp  = df["_mp"].values
rho = df["_rho"].values
ve  = df["_ve"].values
en  = df["_en"].values
sp  = df["_sp"].values

def noise(arr, pct):
    return arr * (1 + np.random.normal(0, pct, size=len(arr)))

# ══════════════════════════════════════════════════════════════════════════════
# GROUP A — THERMODYNAMIC / STRUCTURAL FEATURES
# ══════════════════════════════════════════════════════════════════════════════

# F01 — Melting Point (K)
# Source: Literature / Materials Project mp-* entries
# Physics: Primary UHTM selection criterion; directly controls max service temp
F01 = noise(mp, 0.02)

# F02 — Debye Temperature (K)
# Physics: θ_D ≈ 300 + 0.15*Tm  (Debye-Grüneisen scaling with Tm)
# Relevance: Controls phonon spectrum cutoff; correlates with κ and Cv at high T
F02 = noise(300 + mp * 0.15, 0.04)

# F03 — Cohesive Energy (eV/atom)
# Physics: E_coh = ve*0.82 + en*1.1  (metallic + ionic/covalent bond contributions)
# Relevance: Higher E_coh → greater thermodynamic stability at extreme T
F03 = noise(ve * 0.82 + en * 1.1, 0.05)

# F04 — Formation Enthalpy (kJ/mol)
# Physics: ΔHf = -(en*45 + ve*8)  negative = exothermic = thermodynamically stable
# Relevance: Confirms phase stability; linked to oxidation driving force
F04 = noise(-(en * 45 + ve * 8), 0.05)

# F05 — Lattice Parameter a (Å)
# Physics: a ≈ 3.2 + ρ^(-0.3)*0.5  (inverse power law with density)
# Relevance: Determines interatomic spacing; needed for XRD phase ID and strain calc
F05 = noise(3.2 + (rho ** -0.3) * 0.5, 0.02)

# F06 — Grüneisen Parameter (dimensionless)
# Physics: γ ≈ 0.4 + en*0.3 + ve*0.05  (anharmonicity index)
# Relevance: Links thermal expansion to phonon anharmonicity; key for CTE prediction
F06 = noise(0.4 + en * 0.3 + ve * 0.05, 0.06)

# ══════════════════════════════════════════════════════════════════════════════
# GROUP B — ELECTRONIC STRUCTURE FEATURES
# ══════════════════════════════════════════════════════════════════════════════

# F07 — Band Gap (eV)
# Physics: Eg ≈ max(0, 0.5 - ve*0.05)  — metallic carbides/nitrides ≈ 0
# Relevance: Determines metallic vs semiconductor behaviour at high T; affects emissivity
F07 = np.maximum(0, noise(0.5 - ve * 0.05, 0.20))

# F08 — DOS at Fermi Level (states/eV)
# Physics: N(Ef) ≈ ve*0.55 + 1.2  (more valence e → higher metallic DOS)
# Relevance: Governs electronic heat capacity Cel = γT; thermal radiation absorption
F08 = noise(ve * 0.55 + 1.2, 0.08)

# F09 — Bader Charge Transfer (e⁻)
# Physics: Δq ≈ en*0.6  (Pauling-type charge transfer proportional to ΔEN)
# Relevance: Quantifies ionic vs covalent bond character; predicts oxidation reactivity
F09 = noise(en * 0.6, 0.07)

# F10 — Fermi Velocity (m/s)
# Physics: v_F = 1e6 * (ve/8.5)  (free-electron model scaling)
# Relevance: Links to σ_el and κ_el via Wiedemann-Franz law: κ_el = L₀*σ*T
F10 = noise(1e6 * (ve / 8.5), 0.06)

# F11 — Valence Electron Density (×10²² /cm³)
# Physics: n = ρ*ve*1.8e22  (electrons per unit volume)
# Relevance: Controls metallic bonding strength; shear modulus correlates with n^(1/3)
F11 = noise(rho * ve * 1.8e22, 0.05)

# F12 — Work Function (eV)
# Physics: φ ≈ 3.5 + en*0.8 - ve*0.05  (surface electron escape energy)
# Relevance: Controls thermionic emission at high T; surface reactivity and catalysis
F12 = noise(3.5 + en * 0.8 - ve * 0.05, 0.06)

# ── Build output DataFrame ─────────────────────────────────────────────────────
meta_cols = ["Sample_ID", "Material_System", "Source_Type", "Synthesis_Method", "Crystal_Structure"]
out = df[meta_cols].copy()

out["F01_Melting_Point_K"]          = F01.round(1)
out["F02_Debye_Temperature_K"]      = F02.round(1)
out["F03_Cohesive_Energy_eVatom"]   = F03.round(4)
out["F04_Formation_Enthalpy_kJmol"] = F04.round(3)
out["F05_Lattice_Param_a_Ang"]      = F05.round(4)
out["F06_Gruneisen_Parameter"]      = F06.round(4)
out["F07_Band_Gap_eV"]              = F07.round(4)
out["F08_DOS_at_Fermi_statesEV"]    = F08.round(4)
out["F09_Bader_Charge_Transfer_e"]  = F09.round(4)
out["F10_Fermi_Velocity_ms"]        = F10.round(0)
out["F11_Valence_Electron_Density"] = F11.round(3)
out["F12_Work_Function_eV"]         = F12.round(4)

# ── Save to xlsx ───────────────────────────────────────────────────────────────
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Member1_F01_F12"

NAVY    = "1B2A4A"
GREEN   = "15803D"
LBLUE   = "DBEAFE"
LGREEN  = "DCFCE7"
WHITE   = "FFFFFF"

feature_cols = [c for c in out.columns if c.startswith("F")]

for ci, col in enumerate(out.columns, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if col in meta_cols:
        cell.fill = PatternFill("solid", start_color=NAVY)
    elif col.startswith("F0") or col in ["F10_Fermi_Velocity_ms","F11_Valence_Electron_Density","F12_Work_Function_eV"]:
        cell.fill = PatternFill("solid", start_color="0891B2")   # teal = group A/B
    else:
        cell.fill = PatternFill("solid", start_color=GREEN)

EXP = PatternFill("solid", start_color="EFF6FF")
SYN = PatternFill("solid", start_color="FFFBEB")

for ri, row in out.iterrows():
    bg = EXP if row["Source_Type"] == "experimental" else SYN
    for ci, col in enumerate(out.columns, 1):
        cell = ws.cell(row=ri+2, column=ci, value=row[col])
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col in meta_cols:
            cell.fill = bg

for ci, col in enumerate(out.columns, 1):
    ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) * 0.75)
ws.row_dimensions[1].height = 40
ws.freeze_panes = "F2"

wb.save("Aadi.xlsx")
print(f"Member 1 done. Shape: {out.shape}")
print(f"Features: {feature_cols}")
