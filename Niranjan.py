
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

np.random.seed(42)

# ── Load base file ─────────────────────────────────────────────────────────────
df = pd.read_excel("UHTM_base_200.xlsx", sheet_name="Base Data")

mp  = df["_mp"].values
rho = df["_rho"].values
ve  = df["_ve"].values
en  = df["_en"].values
sp  = df["_sp"].values

def noise(arr, pct):
    return arr * (1 + np.random.normal(0, pct, size=len(arr)))

# Recompute dependent quantities needed for Group H and Targets
# (same seeds as other members because np.random.seed(42) is set above)
youngs_mod   = noise(200 + mp * 0.04 + ve * 15, 0.05)
hardness     = noise(15 + en * 8 + ve * 1.2, 0.06)
kth          = noise(20 + ve * 4 - en * 3, 0.07)
cte          = noise(6.5 + en * 0.8 - ve * 0.3, 0.06)
kic          = noise(2.5 + 1.5 / en, 0.07)
cp           = noise(180 + 800 / rho, 0.05)
rel_dens     = np.minimum(99.9, noise(92 + sp * 0.18, 0.02))
porosity     = np.maximum(0.1, noise(100 - rel_dens, 0.15))
cohesive_e   = noise(ve * 0.82 + en * 1.1, 0.05)
ox_onset     = noise(800 + en * 150 + mp * 0.05, 0.05)
ox_rate      = noise(1e-12 * np.exp(en * 0.8), 0.10)
ox_stability = noise(en * 0.7 + ve * 0.15, 0.07)
grain_size   = noise(2 + (sp ** -0.3) * 10, 0.10)

is_composite = (df["Source_Type"] == "experimental") & (df.index >= 50)

# ══════════════════════════════════════════════════════════════════════════════
# GROUP G — PHASE / COMPOSITE FEATURES
# ══════════════════════════════════════════════════════════════════════════════

# F37 — Phase Stability Index (dimensionless)
# Physics: PSI = Tm/4000 + Ecoh/12  — normalised combined thermodynamic stability
# Relevance: Predicts risk of phase decomposition or polymorphic transition at service T
F37 = noise(mp / 4000 + cohesive_e / 12, 0.06)

# F38 — Secondary Phase Volume Fraction (%)
# Physics: 0% for monolithic; 5–20% for composites (idx-dependent for variety)
# Relevance: Composite reinforcement fraction; tunes toughness and oxidation behaviour
idx_arr = df.index.values
F38 = np.where(is_composite.values,
               noise(5 + (idx_arr % 30) * 0.5, 0.15),
               0.0)

# F39 — Interfacial Energy (J/m²)
# Physics: γ_int = 0.5 + en*0.3  (higher ionic mismatch → higher interfacial energy)
# Relevance: Governs composite phase adhesion; high γ_int → delamination risk
F39 = noise(0.5 + en * 0.3, 0.08)

# F40 — CTE Mismatch (dimensionless index)
# Physics: ΔCTE = |CTE - 5.0| * 0.3  (relative to SiC CTE ~5×10⁻⁶/K)
# Relevance: Drives residual stress and thermal fatigue cracking in composite systems
F40 = noise(np.abs(cte - 5.0) * 0.3, 0.10)

# F41 — Solid Solution Distortion Parameter δ (dimensionless)
# Physics: δ = en*0.12 + (ve%3)*0.05  (Hume-Rothery inspired lattice distortion)
# Relevance: Quantifies solid solution strengthening potential from atomic size mismatch
F41 = noise(en * 0.12 + (ve % 3) * 0.05, 0.08)

# F42 — Wettability Index (dimensionless)
# Physics: W = 1 - bond_ionicity = ve*0.15 / (en + ve*0.15)  (covalent fraction)
# Relevance: Controls sintering kinetics, liquid-phase wetting, and coating adhesion
bond_ionicity = en / (en + ve * 0.15)
F42 = noise(1 - bond_ionicity, 0.07)

# ══════════════════════════════════════════════════════════════════════════════
# GROUP H — ML DESCRIPTOR FEATURES
# ══════════════════════════════════════════════════════════════════════════════

# F43 — Thermal Merit Index (W/kg)
# Physics: TMI = κ / (CTE × ρ)  — Ashby figure of merit for TPS panels
# Relevance: Pareto front descriptor: maximise thermal performance per unit mass
F43 = kth / (cte * rho)

# F44 — Toughness-Stiffness Index (GPa·√GPa)
# Physics: TSI = KIc × √E  — combined crack resistance and stiffness
# Relevance: Captures synergistic crack resistance; cannot be maximised independently
F44 = kic * (youngs_mod ** 0.5)

# F45 — Oxidation Merit Score (dimensionless)
# Physics: OMS = T_ox / (kp_norm + 1)  — onset temperature penalised by rate constant
# Relevance: Directly usable as reward signal in Bayesian optimisation / RL
F45 = ox_onset / (ox_rate * 1e12 + 1)

# F46 — Bond Ionicity Fraction (dimensionless)
# Physics: f_ion = ΔEN / (ΔEN + ve*0.15)  — Pauling ionicity fraction
# Relevance: Single scalar encoding electronic bonding character; governs reactivity
F46 = noise(bond_ionicity, 0.05)

# F47 — Structural Stability Index (dimensionless, 0–1)
# Physics: SSI = (Tm/4000) × (Ecoh/10) × (1 − P/100)
# Relevance: Multi-group composite index; ideal normalised objective for inverse design
F47 = (mp / 4000) * (cohesive_e / 10) * (1 - porosity / 100)

# F48 — Creep Resistance Parameter (dimensionless)
# Physics: CR = (Tm/3000) × (E/400) × (1/grain_size)^0.3
# Relevance: Estimates high-temperature creep resistance via grain size and stiffness
F48 = noise((mp / 3000) * (youngs_mod / 400) * (1 / grain_size) ** 0.3, 0.08)

# ══════════════════════════════════════════════════════════════════════════════
# TARGET VARIABLES
# ══════════════════════════════════════════════════════════════════════════════

# T1 — Flexural Strength (MPa)
# Multi-factor: stiff + hard + dense → strong
T1 = noise(300 + youngs_mod * 0.8 + hardness * 12 - porosity * 15, 0.05)
T1 = np.maximum(50, T1)

# T2 — Oxidation Resistance Score (0–10)
# Weighted: onset(×3) + rate(×2) + stability(×2) + CTE_match(×1)
T2 = np.minimum(10, np.maximum(0, noise(
    3 * (ox_onset / 1500) + 2 * (1 / (ox_rate * 1e10 + 0.1)) +
    2 * ox_stability + 1 * (1 - F40), 0.05)))

# T3 — Thermal Shock Resistance (cycles to failure)
# Fracture toughness + low CTE + low mismatch → more cycles
T3 = np.maximum(1, noise(
    20 + kic * 8 + (1 / cte) * 0.5 - F40 * 5, 0.08).astype(int))

# ── Build output DataFrame ─────────────────────────────────────────────────────
meta_cols = ["Sample_ID", "Material_System", "Source_Type", "Synthesis_Method", "Crystal_Structure"]
out = df[meta_cols].copy()

out["F37_Phase_Stability_Index"]       = F37.round(5)
out["F38_Secondary_Phase_vol_pct"]     = F38.round(2)
out["F39_Interfacial_Energy_Jm2"]      = F39.round(4)
out["F40_CTE_Mismatch_Index"]          = F40.round(5)
out["F41_Solid_Solution_Delta"]        = F41.round(5)
out["F42_Wettability_Index"]           = F42.round(4)
out["F43_Thermal_Merit_Index_Wkg"]     = F43.round(4)
out["F44_Toughness_Stiffness_Index"]   = F44.round(4)
out["F45_Oxidation_Merit_Score"]       = F45.round(4)
out["F46_Bond_Ionicity"]               = F46.round(5)
out["F47_Structural_Stability_Index"]  = F47.round(6)
out["F48_Creep_Resistance_Param"]      = F48.round(5)
out["T1_Flexural_Strength_MPa"]        = T1.round(1)
out["T2_Oxidation_Resistance_Score"]   = T2.round(3)
out["T3_Thermal_Shock_Cycles"]         = T3.astype(int)

# ── Save to xlsx ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Member4_F37_F48_Targets"

NAVY   = "1B2A4A"
INDIGO = "3730A3"
TEAL   = "0F766E"
RED    = "991B1B"
WHITE  = "FFFFFF"

PHASE_COLS  = ["F37_Phase_Stability_Index","F38_Secondary_Phase_vol_pct","F39_Interfacial_Energy_Jm2",
               "F40_CTE_Mismatch_Index","F41_Solid_Solution_Delta","F42_Wettability_Index"]
ML_COLS     = ["F43_Thermal_Merit_Index_Wkg","F44_Toughness_Stiffness_Index","F45_Oxidation_Merit_Score",
               "F46_Bond_Ionicity","F47_Structural_Stability_Index","F48_Creep_Resistance_Param"]
TARGET_COLS = ["T1_Flexural_Strength_MPa","T2_Oxidation_Resistance_Score","T3_Thermal_Shock_Cycles"]

for ci, col in enumerate(out.columns, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if col in meta_cols:
        cell.fill = PatternFill("solid", start_color=NAVY)
    elif col in PHASE_COLS:
        cell.fill = PatternFill("solid", start_color=INDIGO)
    elif col in ML_COLS:
        cell.fill = PatternFill("solid", start_color=TEAL)
    else:
        cell.fill = PatternFill("solid", start_color=RED)

EXP = PatternFill("solid", start_color="EFF6FF")
SYN = PatternFill("solid", start_color="FFFBEB")

for ri, row in out.iterrows():
    bg = EXP if row["Source_Type"] == "experimental" else SYN
    for ci, col in enumerate(out.columns, 1):
        cell = ws.cell(row=ri+2, column=ci, value=row[col])
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col in meta_cols:
            cell.fill = bg

for ci, col in enumerate(out.columns, 1):
    ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) * 0.75)
ws.row_dimensions[1].height = 40
ws.freeze_panes = "F2"

wb.save("Niranjan.xlsx")
print(f"Member 4 done. Shape: {out.shape}")
print(f"Features: {[c for c in out.columns if c.startswith('F')]}")
print(f"Targets:  {[c for c in out.columns if c.startswith('T')]}")
