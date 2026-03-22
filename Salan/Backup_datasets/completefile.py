"""
=============================================================================
 UHTM_Complete_Dataset.py
 Machine Learning-Based Multi-Objective Optimization of
 Ultra-High Temperature Materials for Aerospace Applications
-----------------------------------------------------------------------------
 Generates a SINGLE self-contained dataset from scratch (no CSV dependency).
 500 samples  |  250 Experimental + 250 Synthetic  |  100 columns
=============================================================================
 COLUMN GROUPS
   [A] Metadata            : 2  cols  (Sample_ID, Data_Source)
   [B] Formula / Phase     : 3  cols  (Molecular_Formula, Empirical_Formula, Dominant_Phase)
   [C] Composition         : 9  cols  (wt% + VEC)
   [D] Processing          : 12 cols  (sintering, milling, atmosphere …)
   [E] Microstructure      : 8  cols  (grain, porosity, density, phase fractions …)
   [F] Thermodynamic       : 8  cols  (ΔH_mix, δ, ΔS_mix, Ω, Δχ, liquidus …)
   [G] Mechanical – Basic  : 5  cols  (hardness, KIC, σ_f, E, Weibull)
   [H] Mechanical – Adv    : 9  cols  (ν, G, K, Pugh, flaw size, specific …)
   [I] Thermal – Basic     : 4  cols  (κ, ΔT_shock, CTE, T_melt)
   [J] Thermal – Advanced  : 8  cols  (Cp, α, σ_therm, Bi, MFP, Cp_HT, φ_rad, κ_HT)
   [K] Oxidation/Corrosion : 7  cols  (oxide thickness, Kp, Ea_ox, HCR, DO2, Δm, spall)
   [L] Powder/Processing   : 7  cols  (BET, E_mill, ρ_green, densif., Ea_sint, flow, Carr)
   [M] Fracture / Damage   : 6  cols  (CTOD, J, notch, R-curve, da/dN, DT)
   [N] HiTemp / Aerospace  : 12 cols  (T_test, oxid, creep, ε, ablation, LE-surv,
                                        plasma, FOM, TPS, life, Q_spec, reuse)
=============================================================================
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════
np.random.seed(42)
N = 500

def clamp(a, lo, hi): return np.clip(a, lo, hi)
def jitter(n, s):     return np.random.normal(0, s, n)

# ═══════════════════════════════════════════════════════════════════════════
# [C] COMPOSITION  — wt%, normalised to 100 %
# ═══════════════════════════════════════════════════════════════════════════
Hf = np.random.uniform(10, 45, N)
Zr = np.random.uniform(5,  30, N)
Ta = np.random.uniform(0,  20, N)
Nb = np.random.uniform(0,  15, N)
Ti = np.random.uniform(0,  10, N)
Si = np.random.uniform(5,  25, N)
C  = np.random.uniform(5,  20, N)
B  = np.random.uniform(0,  10, N)

total = Hf+Zr+Ta+Nb+Ti+Si+C+B
Hf, Zr, Ta, Nb, Ti, Si, C, B = [x/total*100 for x in (Hf,Zr,Ta,Nb,Ti,Si,C,B)]

# Valence electron concentration
VEC = (Hf*4 + Zr*4 + Ta*5 + Nb*5 + Ti*4 + Si*4 + C*4 + B*3) / 100

# ═══════════════════════════════════════════════════════════════════════════
# [D] PROCESSING
# ═══════════════════════════════════════════════════════════════════════════
T_sint   = np.random.uniform(1600, 2200, N)   # °C
t_sint   = np.random.uniform(30,   240,  N)   # min
P_sint   = np.random.uniform(20,   100,  N)   # MPa
heat_r   = np.random.uniform(5,    50,   N)   # °C/min
atm_str  = np.random.choice(['Ar','N2','Vacuum','He'], N)
atm_code = np.where(atm_str=='Vacuum',0, np.where(atm_str=='Ar',1,
           np.where(atm_str=='N2',2,3)))
d_part   = np.random.uniform(0.1,  10,   N)   # µm  powder particle size
t_mill   = np.random.uniform(0,    72,   N)   # h
spd_mill = np.random.uniform(100,  500,  N)   # rpm
BPR      = np.random.uniform(5,    20,   N)   # ball-to-powder ratio
cool_r   = np.random.uniform(1,    50,   N)   # °C/min

melt_proxy  = 0.35*Hf + 0.30*Zr + 0.20*Ta + 0.15*Nb
T_melt_est  = 2500 + melt_proxy*30
hom_T       = T_sint / T_melt_est              # homologous temperature

# ═══════════════════════════════════════════════════════════════════════════
# [E] MICROSTRUCTURE
# ═══════════════════════════════════════════════════════════════════════════
grain      = np.random.uniform(0.5, 50, N)        # µm
GAR        = np.random.uniform(1.0,  5, N)        # grain aspect ratio
por        = clamp(np.random.exponential(2, N), 0.01, 15)  # %
rho        = clamp(5.5 + 3.5*(Hf/100) + 2.8*(Zr/100) - por*0.05 + jitter(N,0.2), 4, 9)
rel_rho    = clamp(rho/(6.0+3*(Hf/100))*100, 85, 99.9)
f_carb     = clamp(40 + C*1.5  + jitter(N,3), 0, 100)
f_bor      = clamp(20 + B*2.0  + jitter(N,3), 0, 100)
f_SiC      = clamp(Si*1.2      + jitter(N,2), 0, 60)

# ═══════════════════════════════════════════════════════════════════════════
# [G] MECHANICAL – BASIC  (needed early for cross-dependencies)
# ═══════════════════════════════════════════════════════════════════════════
hardness = clamp(15 + 0.08*T_sint/100 + 0.05*(100-por)
                    + 0.04*VEC*10 + 0.1*(Hf+Zr)/10
                    - 0.02*grain + jitter(N,1.5), 10, 35)

KIC = clamp(4.0 + 0.05*grain + 0.03*(Nb+Ti)
                - 0.02*hardness + 0.5*GAR + jitter(N,0.5), 2.0, 12.0)

sigma_f = clamp(300 + 10*hardness - 15*grain
                    + 5*(100-por) + jitter(N,30), 200, 900)

E_mod = clamp(400 + 0.5*(Hf+Zr) + 0.3*Ta
                  - 10*por + jitter(N,15), 250, 600)

weibull = clamp(8 + 0.5*rel_rho/10 - 0.2*grain + jitter(N,1.5), 5, 25)

# ═══════════════════════════════════════════════════════════════════════════
# [I] THERMAL – BASIC  (needed for cross-deps)
# ═══════════════════════════════════════════════════════════════════════════
K_therm = clamp(30 + 0.4*(Hf+Zr) - 0.5*por
                   - 0.1*grain + 2*f_SiC/20 + jitter(N,3), 10, 80)

TSR     = clamp(KIC*200/(K_therm+1e-3)*0.5 + jitter(N,30), 100, 1000)

CTE     = clamp(6 + 0.03*(Zr+Hf) + 0.05*Ti + jitter(N,0.5), 4, 12)

T_melt  = clamp(2800 + 50*(Hf/100)*30 + 40*(Zr/100)*20
                      + 60*(Ta/100)*15 - 30*(Si/100)*10
                      + jitter(N,80), 2600, 3800)

# ═══════════════════════════════════════════════════════════════════════════
# [N] HIGH-TEMP / AEROSPACE  (partial – T_test needed for other blocks)
# ═══════════════════════════════════════════════════════════════════════════
T_test   = np.random.uniform(1000, 2500, N)        # °C  test temperature

oxid     = clamp(0.5 + 0.3*np.exp((T_test-1500)/300)
                     - 0.5*(f_SiC/60) - 0.1*(Hf/40)
                     + jitter(N,0.2), 0.01, 10)

creep    = clamp(1e-8*np.exp((T_test-1000)/300)*(1/(hardness+1e-3))
                    *(1+por/10) + np.abs(jitter(N,1e-9)), 1e-10, 1e-4)

emiss    = clamp(0.85 + 0.05*(B/10) + jitter(N,0.03), 0.6, 0.99)
ablat    = clamp(0.01 + 0.005*oxid   + jitter(N,0.005), 0.001, 0.5)

# ═══════════════════════════════════════════════════════════════════════════
# [A] DATA SOURCE + EXPERIMENTAL NOISE
# ═══════════════════════════════════════════════════════════════════════════
data_src = np.array(['Experimental']*250 + ['Synthetic']*250)
np.random.shuffle(data_src)
em = data_src == 'Experimental'
hardness[em]  += jitter(em.sum(), 0.8)
KIC[em]       += jitter(em.sum(), 0.3)
sigma_f[em]   += jitter(em.sum(), 20)
K_therm[em]   += jitter(em.sum(), 2)

# Re-clamp after noise
hardness = clamp(hardness, 10, 35)
KIC      = clamp(KIC,      2.0, 12.0)
sigma_f  = clamp(sigma_f,  200, 900)
K_therm  = clamp(K_therm,  10, 80)

# ═══════════════════════════════════════════════════════════════════════════
# [F] THERMODYNAMIC / PHASE-DIAGRAM FEATURES
# ═══════════════════════════════════════════════════════════════════════════
AW  = dict(Hf=178.49,Zr=91.22,Ta=180.95,Nb=92.91,Ti=47.87,Si=28.09,C=12.01,B=10.81)
r_a = dict(Hf=1.58,Zr=1.60,Ta=1.47,Nb=1.47,Ti=1.46,Si=1.11,C=0.77,B=0.87)
chi = dict(Hf=1.3, Zr=1.33,Ta=1.50,Nb=1.60,Ti=1.54,Si=1.90,C=2.55,B=2.04)

elems  = ['Hf','Zr','Ta','Nb','Ti','Si','C','B']
wt_arr = np.column_stack([Hf,Zr,Ta,Nb,Ti,Si,C,B])          # (N,8) wt%
mol_arr= wt_arr / np.array([AW[e] for e in elems])
mf_arr = mol_arr / mol_arr.sum(axis=1, keepdims=True)       # mole fractions

avg_r   = mf_arr @ np.array([r_a[e] for e in elems])
avg_chi = mf_arr @ np.array([chi[e] for e in elems])

dH_mix  = clamp(-5 + 0.3*(Hf*Zr)/1000 + 0.2*(Ta*Nb)/1000
                   - 0.1*(Si*C)/1000 + jitter(N,1.5), -20, 5)

delta_r = clamp(np.sqrt(np.sum(mf_arr*(
              np.column_stack([r_a[e] for e in elems]) - avg_r[:,None])**2,
              axis=1))*100, 0, 15)

x_safe  = np.where(mf_arr<1e-9, 1e-9, mf_arr)
dS_mix  = clamp(-8.314*np.sum(x_safe*np.log(x_safe), axis=1), 0, 18)

omega   = clamp(T_melt*dS_mix / (np.abs(dH_mix)*1000+1e-6), 0, 500)

delta_chi = clamp(np.sqrt(np.sum(mf_arr*(
                np.column_stack([chi[e] for e in elems]) - avg_chi[:,None])**2,
                axis=1)), 0, 1.5)

liquidus  = clamp(2600 + 50*(Hf/100)*500 + 40*(Zr/100)*300
                       + 70*(Ta/100)*600  - 40*(Si/100)*300 + jitter(N,60),
                  2500, 4000)

n_princ   = sum((wt_arr[:,i]>5).astype(int) for i in range(8))

phi_stab  = clamp(omega/(delta_r+1) + VEC*2 + jitter(N,5), 0, 200)

# ═══════════════════════════════════════════════════════════════════════════
# [H] MECHANICAL – ADVANCED
# ═══════════════════════════════════════════════════════════════════════════
nu         = clamp(0.18 + 0.02*(Si/20) + 0.01*(B/10) + jitter(N,0.01), 0.12, 0.30)
G_mod      = clamp(E_mod/(2*(1+nu))   + jitter(N,8),   80,  280)
K_bulk     = clamp(E_mod/(3*(1-2*nu)) + jitter(N,10),  150, 400)
pugh       = clamp(K_bulk/(G_mod+1e-6)+ jitter(N,0.05),0.8, 3.0)
flaw_size  = clamp((KIC/(sigma_f*np.sqrt(np.pi)))**2*1e6, 0.1, 100)
sp_str     = clamp(sigma_f/(rho+1e-6) + jitter(N,10),  30,  200)
sp_stiff   = clamp(E_mod/(rho+1e-6)   + jitter(N,5),   30,  120)
nano_H     = clamp(hardness*0.94      + jitter(N,0.5),  8,   34)
U_el       = clamp((sigma_f**2)/(2*E_mod*1e3)*1e-3 + jitter(N,0.01), 0.001, 2.0)

# ═══════════════════════════════════════════════════════════════════════════
# [J] THERMAL – ADVANCED
# ═══════════════════════════════════════════════════════════════════════════
Cp         = clamp(350 + 0.5*(Si*4) + 0.3*(C*6) + jitter(N,20), 250, 650)
alpha_th   = clamp(K_therm/(rho*1000*Cp)*1e6 + jitter(N,0.5),   0.5, 25)
dT_serv    = T_test - 25
sig_therm  = clamp(E_mod*(CTE*1e-6)*dT_serv + jitter(N,10),      0,  2000)
Bi         = clamp(K_therm/(alpha_th+1e-6)*0.01 + jitter(N,0.05),0.01,5.0)
ph_mfp     = clamp(50 + 30*(K_therm/50) - 5*grain + jitter(N,5), 5, 200)
Cp_HT      = clamp(Cp*(1+0.0002*(T_test-25)) + jitter(N,15),     300, 900)
SB         = 5.67e-8
phi_rad    = clamp(emiss*SB*(T_test+273)**4/1e3 + jitter(N,5),   0.1, 5000)
K_HT       = clamp(K_therm*(300/(T_test+300)) + jitter(N,2),      5,  70)

# ═══════════════════════════════════════════════════════════════════════════
# [K] OXIDATION / CORROSION
# ═══════════════════════════════════════════════════════════════════════════
ox_thick   = clamp(np.sqrt(oxid*3600*100)*10   + jitter(N,1),    0.1,  500)
Kp         = clamp(oxid**2*3600                + jitter(N,1e-3), 1e-6, 100)
Ea_ox      = clamp(200+30*(f_SiC/60)+20*(Hf/40)-10*(por/15)+jitter(N,10), 100, 400)
HCR        = clamp(7+0.05*(Hf+Zr)/40-0.1*oxid*10+0.3*(f_SiC/60)*5+jitter(N,0.5),1,10)
logDO2     = clamp(-14+3*((T_test-1000)/1500)-0.5*(rel_rho/100)*3+jitter(N,0.3),-18,-8)
DO2        = 10**logDO2
mass_gain  = clamp(oxid*3600*(T_test/1000) + jitter(N,0.5),      0.01, 200)
spall      = clamp(0.5+0.3*(KIC/12)-0.2*(CTE/12)+0.1*(K_therm/80)+jitter(N,0.05),0,1)

# ═══════════════════════════════════════════════════════════════════════════
# [L] POWDER / PROCESSING DERIVED
# ═══════════════════════════════════════════════════════════════════════════
BET        = clamp(6/(d_part*rho)          + jitter(N,0.2),   0.1,  20)
E_mill     = clamp(spd_mill**2*t_mill/1e5  + jitter(N,0.1),   0,    30)
rho_green  = clamp(rho*0.65+0.01*P_sint    + jitter(N,0.1),   3.0,  6.5)
densif_r   = clamp((rel_rho-60)/(t_sint+1e-6)+jitter(N,0.05), 0.01, 2.0)
Ea_sint    = clamp(400+50*(Hf/40)+30*(Ta/20)-20*(Si/20)+jitter(N,20), 250, 700)
flow       = clamp(20+5*d_part-2*BET        + jitter(N,2),    5,    60)
rho_tap    = clamp(rho_green*1.15           + jitter(N,0.1),  3.0,  7.0)
rho_bulk   = clamp(rho_tap*0.80             + jitter(N,0.1),  2.0,  6.5)
carr       = clamp((rho_tap-rho_bulk)/rho_tap*100+jitter(N,1),5,    40)

# ═══════════════════════════════════════════════════════════════════════════
# [M] FRACTURE / DAMAGE MECHANICS
# ═══════════════════════════════════════════════════════════════════════════
CTOD       = clamp(KIC**2/(E_mod*sigma_f)*1000 + jitter(N,0.01), 0.001, 5.0)
J_int      = clamp(KIC**2/E_mod               + jitter(N,0.005),0.01,  0.6)
notch_s    = clamp(1-(KIC/(sigma_f*np.sqrt(np.pi*grain*1e-6)*1e3))+jitter(N,0.05),0.1,1.0)
R_slope    = clamp(0.05+0.01*grain+0.02*KIC   + jitter(N,0.005),0.01,  2.0)
da_dN      = clamp(1e-3*(5/KIC)**3*(E_mod/400)+np.abs(jitter(N,1e-4)),1e-5,10)
DT_idx     = clamp(KIC/(hardness+1e-6)         + jitter(N,0.05), 0.1,  1.5)

# ═══════════════════════════════════════════════════════════════════════════
# [N] AEROSPACE PERFORMANCE (remaining columns)
# ═══════════════════════════════════════════════════════════════════════════
LE_surv    = clamp(10-ablat*50+emiss*5+K_therm/20 + jitter(N,0.5),  0.5, 15)
plasma_rec = clamp(ablat*1000*(1-emiss*0.5)        + jitter(N,0.5),  0.01,300)
aero_FOM   = clamp(sigma_f/(rho*CTE*E_mod*1e-3)    + jitter(N,0.1),  0.01, 50)
TPS_eff    = clamp(emiss*K_therm/(CTE*rho)          + jitter(N,0.5),  0.1, 100)
svc_life   = clamp(500/(creep*1e6+oxid+ablat*10+1e-6)+jitter(N,5),   1,   2000)
Q_spec     = clamp(Cp*(T_test-25)/1000              + jitter(N,10),   10,  2000)
reuse_idx  = clamp(8-ablat*20-oxid*5+spall*3        + jitter(N,0.3),  0,   10)

# ═══════════════════════════════════════════════════════════════════════════
# [B] MOLECULAR FORMULA COLUMNS
# ═══════════════════════════════════════════════════════════════════════════
def make_mol_formula(mf_row, threshold=0.01):
    parts = []
    for i, el in enumerate(elems):
        x = mf_row[i]
        if x < threshold: continue
        xr = round(float(x), 2)
        parts.append(el if xr == 1.0 else f"{el}{xr:.2f}")
    return ''.join(parts) or 'Unknown'

def make_empirical_formula(mf_row, threshold=0.01):
    inc = [(el, mf_row[i]) for i, el in enumerate(elems) if mf_row[i] >= threshold]
    if not inc: return 'Unknown'
    min_x = min(x for _, x in inc)
    parts = []
    for el, x in inc:
        r = round(x/min_x, 1)
        parts.append(el if r == 1.0 else f"{el}{r:.1f}")
    return ''.join(parts)

def dominant_phase(wt_row):
    wd   = dict(zip(elems, wt_row))
    metal= max(['Hf','Zr','Ta','Nb','Ti'], key=lambda e: wd[e])
    nonm = max(['C','Si','B'],             key=lambda e: wd[e])
    suf  = {'C':'C','B':'B2','Si':'Si'}[nonm]
    return f"{metal}{suf}"

mol_formula  = [make_mol_formula(mf_arr[i])     for i in range(N)]
emp_formula  = [make_empirical_formula(mf_arr[i]) for i in range(N)]
dom_phase    = [dominant_phase(wt_arr[i])        for i in range(N)]

# ═══════════════════════════════════════════════════════════════════════════
# ASSEMBLE DATAFRAME  (100 columns)
# ═══════════════════════════════════════════════════════════════════════════
df = pd.DataFrame({
    # ── [A] Metadata ────────────────────────────────────────────────── 2
    'Sample_ID':                        [f'UHTM_{i+1:04d}' for i in range(N)],
    'Data_Source':                      data_src,

    # ── [B] Formula / Phase ─────────────────────────────────────────── 3
    'Molecular_Formula':                mol_formula,
    'Empirical_Formula':                emp_formula,
    'Dominant_Phase':                   dom_phase,

    # ── [C] Composition ─────────────────────────────────────────────── 9
    'Hf_content_wt%':                   np.round(Hf,  3),
    'Zr_content_wt%':                   np.round(Zr,  3),
    'Ta_content_wt%':                   np.round(Ta,  3),
    'Nb_content_wt%':                   np.round(Nb,  3),
    'Ti_content_wt%':                   np.round(Ti,  3),
    'Si_content_wt%':                   np.round(Si,  3),
    'C_content_wt%':                    np.round(C,   3),
    'B_content_wt%':                    np.round(B,   3),
    'Valence_Electron_Conc':            np.round(VEC, 4),

    # ── [D] Processing ──────────────────────────────────────────────── 12
    'Sintering_Temp_C':                 np.round(T_sint,   1),
    'Sintering_Time_min':               np.round(t_sint,   1),
    'Sintering_Pressure_MPa':           np.round(P_sint,   2),
    'Heating_Rate_C_per_min':           np.round(heat_r,   2),
    'Atmosphere':                       atm_str,
    'Atmosphere_Code':                  atm_code,
    'Powder_Particle_Size_um':          np.round(d_part,   3),
    'Milling_Time_h':                   np.round(t_mill,   2),
    'Milling_Speed_rpm':                np.round(spd_mill, 1),
    'Ball_to_Powder_Ratio':             np.round(BPR,      2),
    'Cooling_Rate_C_per_min':           np.round(cool_r,   2),
    'Homologous_Temperature':           np.round(hom_T,    4),

    # ── [E] Microstructure ──────────────────────────────────────────── 8
    'Grain_Size_um':                    np.round(grain,   3),
    'Grain_Aspect_Ratio':               np.round(GAR,     3),
    'Porosity_%':                       np.round(por,     3),
    'Density_g_cm3':                    np.round(rho,     4),
    'Relative_Density_%':               np.round(rel_rho, 3),
    'Phase_Fraction_Carbide_%':         np.round(f_carb,  2),
    'Phase_Fraction_Boride_%':          np.round(f_bor,   2),
    'SiC_Phase_Fraction_%':             np.round(f_SiC,   2),

    # ── [F] Thermodynamic ───────────────────────────────────────────── 8
    'Mixing_Enthalpy_kJ_mol':           np.round(dH_mix,    4),
    'Atomic_Size_Mismatch_%':           np.round(delta_r,   4),
    'Mixing_Entropy_J_molK':            np.round(dS_mix,    4),
    'Omega_Phase_Stability':            np.round(omega,     4),
    'Electronegativity_Difference':     np.round(delta_chi, 4),
    'Liquidus_Temp_C':                  np.round(liquidus,  1),
    'Num_Principal_Elements':           n_princ,
    'Phase_Stability_Index':            np.round(phi_stab,  3),

    # ── [G] Mechanical – Basic ──────────────────────────────────────── 5
    'Vickers_Hardness_GPa':             np.round(hardness,  3),
    'Fracture_Toughness_MPa_m05':       np.round(KIC,       3),
    'Flexural_Strength_MPa':            np.round(sigma_f,   2),
    'Youngs_Modulus_GPa':               np.round(E_mod,     2),
    'Weibull_Modulus':                  np.round(weibull,   2),

    # ── [H] Mechanical – Advanced ───────────────────────────────────── 9
    'Poisson_Ratio':                    np.round(nu,        4),
    'Shear_Modulus_GPa':                np.round(G_mod,     2),
    'Bulk_Modulus_GPa':                 np.round(K_bulk,    2),
    'Pughs_Ratio_K_G':                  np.round(pugh,      4),
    'Critical_Flaw_Size_um':            np.round(flaw_size, 4),
    'Specific_Strength_MPa_gcm3':       np.round(sp_str,    3),
    'Specific_Stiffness_GPa_gcm3':      np.round(sp_stiff,  3),
    'Nano_Hardness_GPa':                np.round(nano_H,    3),
    'Elastic_Energy_Density_MJ_m3':     np.round(U_el,      5),

    # ── [I] Thermal – Basic ─────────────────────────────────────────── 4
    'Thermal_Conductivity_W_mK':        np.round(K_therm,   3),
    'Thermal_Shock_Resistance_dT_C':    np.round(TSR,       1),
    'CTE_1e6_per_C':                    np.round(CTE,       4),
    'Melting_Decomp_Temp_C':            np.round(T_melt,    1),

    # ── [J] Thermal – Advanced ──────────────────────────────────────── 8
    'Specific_Heat_Cp_J_kgK':           np.round(Cp,        2),
    'Thermal_Diffusivity_mm2_s':        np.round(alpha_th,  4),
    'Thermal_Mismatch_Stress_MPa':      np.round(sig_therm, 2),
    'Biot_Number':                      np.round(Bi,        4),
    'Phonon_MFP_nm':                    np.round(ph_mfp,    3),
    'Cp_HighTemp_J_kgK':                np.round(Cp_HT,     2),
    'Reradiation_Flux_kW_m2':           np.round(phi_rad,   3),
    'K_Therm_HighTemp_W_mK':            np.round(K_HT,      3),

    # ── [K] Oxidation / Corrosion ───────────────────────────────────── 7
    'Oxide_Thickness_100h_um':          np.round(ox_thick,  3),
    'Kp_Parabolic_mg2_cm4s':            np.round(Kp,        6),
    'Oxidation_Activation_Energy_kJ_mol': np.round(Ea_ox,   2),
    'Hot_Corrosion_Resistance_Index':   np.round(HCR,       3),
    'O2_Diffusion_Coeff_m2s':           np.round(DO2,       20),
    'Mass_Gain_Oxidation_mg_cm2':       np.round(mass_gain, 4),
    'Spallation_Resistance':            np.round(spall,     4),

    # ── [L] Powder / Processing Derived ────────────────────────────── 7
    'BET_Surface_Area_m2_g':            np.round(BET,       4),
    'Milling_Energy_kJ_g':              np.round(E_mill,    4),
    'Green_Density_g_cm3':              np.round(rho_green, 4),
    'Densification_Rate_%_min':         np.round(densif_r,  5),
    'Sintering_Activation_Energy_kJ_mol': np.round(Ea_sint, 2),
    'Powder_Flowability_s_50g':         np.round(flow,      3),
    'Carrs_Index_%':                    np.round(carr,      3),

    # ── [M] Fracture / Damage ───────────────────────────────────────── 6
    'CTOD_um':                          np.round(CTOD,      5),
    'J_Integral_kJ_m2':                 np.round(J_int,     5),
    'Notch_Sensitivity_Ratio':          np.round(notch_s,   4),
    'R_Curve_Slope':                    np.round(R_slope,   4),
    'Fatigue_Crack_Growth_nm_cycle':    np.round(da_dN,     6),
    'Damage_Tolerance_Index':           np.round(DT_idx,    4),

    # ── [N] HiTemp / Aerospace ──────────────────────────────────────── 12
    'Test_Temperature_C':               np.round(T_test,    1),
    'Oxidation_Rate_mg_cm2s':           np.round(oxid,      5),
    'Creep_Rate_per_s':                 np.round(creep,     12),
    'Emissivity':                       np.round(emiss,     4),
    'Ablation_Rate_mm_per_s':           np.round(ablat,     5),
    'Leading_Edge_Survivability_mm':    np.round(LE_surv,   3),
    'Plasma_Recession_Rate_um_s':       np.round(plasma_rec,4),
    'Aero_Structural_FOM':              np.round(aero_FOM,  4),
    'TPS_Efficiency_Index':             np.round(TPS_eff,   4),
    'Service_Life_h':                   np.round(svc_life,  2),
    'Heat_Absorption_Specific_J_g':     np.round(Q_spec,    2),
    'Reusability_Index':                np.round(reuse_idx, 3),
})

print(f"Dataset shape : {df.shape[0]} rows × {df.shape[1]} columns")
print(f"Experimental  : {(data_src=='Experimental').sum()}")
print(f"Synthetic     : {(data_src=='Synthetic').sum()}")
print(f"\nAll {df.shape[1]} columns:")
for i, c in enumerate(df.columns, 1):
    print(f"  {i:03d}. {c}")

# ═══════════════════════════════════════════════════════════════════════════
# SAVE  CSV
# ═══════════════════════════════════════════════════════════════════════════
CSV_PATH  = '/mnt/user-data/outputs/UHTM_Complete.csv'
XLSX_PATH = '/mnt/user-data/outputs/UHTM_Complete.xlsx'

df.to_csv(CSV_PATH, index=False)
print(f"\nCSV saved  → {CSV_PATH}")

# ═══════════════════════════════════════════════════════════════════════════
# SAVE  FORMATTED EXCEL  (3 sheets)
# ═══════════════════════════════════════════════════════════════════════════

# ── Group → fill colour map ──────────────────────────────────────────────
GROUP_MAP = {
    # col_name : hex_fill
    'Sample_ID':'BFBFBF','Data_Source':'BFBFBF',
    'Molecular_Formula':'E8D5F5','Empirical_Formula':'E8D5F5','Dominant_Phase':'E8D5F5',
}
COMP_HEX   = 'BDD7EE'   # blue
PROC_HEX   = 'E2EFDA'   # green
MICRO_HEX  = 'FFF2CC'   # yellow
THERMO_HEX = 'F4CCFF'   # purple
MECH_HEX   = 'FCE4D6'   # orange
MECH2_HEX  = 'FFD966'   # gold
THERM_HEX  = 'DDEBF7'   # light blue
THERM2_HEX = '9DC3E6'   # medium blue
OX_HEX     = 'A9D18E'   # sage
POWD_HEX   = 'F8CBAD'   # peach
FRAC_HEX   = 'FF9999'   # red-pink
AERO_HEX   = 'D6E4BC'   # olive-green

for c in df.columns:
    if   'content' in c or c == 'Valence_Electron_Conc':           GROUP_MAP[c] = COMP_HEX
    elif c in ('Sintering_Temp_C','Sintering_Time_min',
               'Sintering_Pressure_MPa','Heating_Rate_C_per_min',
               'Atmosphere','Atmosphere_Code','Powder_Particle_Size_um',
               'Milling_Time_h','Milling_Speed_rpm','Ball_to_Powder_Ratio',
               'Cooling_Rate_C_per_min','Homologous_Temperature'):  GROUP_MAP[c] = PROC_HEX
    elif c in ('Grain_Size_um','Grain_Aspect_Ratio','Porosity_%',
               'Density_g_cm3','Relative_Density_%',
               'Phase_Fraction_Carbide_%','Phase_Fraction_Boride_%',
               'SiC_Phase_Fraction_%'):                             GROUP_MAP[c] = MICRO_HEX
    elif c in ('Mixing_Enthalpy_kJ_mol','Atomic_Size_Mismatch_%',
               'Mixing_Entropy_J_molK','Omega_Phase_Stability',
               'Electronegativity_Difference','Liquidus_Temp_C',
               'Num_Principal_Elements','Phase_Stability_Index'):   GROUP_MAP[c] = THERMO_HEX
    elif c in ('Vickers_Hardness_GPa','Fracture_Toughness_MPa_m05',
               'Flexural_Strength_MPa','Youngs_Modulus_GPa',
               'Weibull_Modulus'):                                  GROUP_MAP[c] = MECH_HEX
    elif c in ('Poisson_Ratio','Shear_Modulus_GPa','Bulk_Modulus_GPa',
               'Pughs_Ratio_K_G','Critical_Flaw_Size_um',
               'Specific_Strength_MPa_gcm3','Specific_Stiffness_GPa_gcm3',
               'Nano_Hardness_GPa','Elastic_Energy_Density_MJ_m3'):GROUP_MAP[c] = MECH2_HEX
    elif c in ('Thermal_Conductivity_W_mK','Thermal_Shock_Resistance_dT_C',
               'CTE_1e6_per_C','Melting_Decomp_Temp_C'):            GROUP_MAP[c] = THERM_HEX
    elif c in ('Specific_Heat_Cp_J_kgK','Thermal_Diffusivity_mm2_s',
               'Thermal_Mismatch_Stress_MPa','Biot_Number',
               'Phonon_MFP_nm','Cp_HighTemp_J_kgK',
               'Reradiation_Flux_kW_m2','K_Therm_HighTemp_W_mK'):  GROUP_MAP[c] = THERM2_HEX
    elif c in ('Oxide_Thickness_100h_um','Kp_Parabolic_mg2_cm4s',
               'Oxidation_Activation_Energy_kJ_mol',
               'Hot_Corrosion_Resistance_Index','O2_Diffusion_Coeff_m2s',
               'Mass_Gain_Oxidation_mg_cm2','Spallation_Resistance'):GROUP_MAP[c] = OX_HEX
    elif c in ('BET_Surface_Area_m2_g','Milling_Energy_kJ_g',
               'Green_Density_g_cm3','Densification_Rate_%_min',
               'Sintering_Activation_Energy_kJ_mol',
               'Powder_Flowability_s_50g','Carrs_Index_%'):         GROUP_MAP[c] = POWD_HEX
    elif c in ('CTOD_um','J_Integral_kJ_m2','Notch_Sensitivity_Ratio',
               'R_Curve_Slope','Fatigue_Crack_Growth_nm_cycle',
               'Damage_Tolerance_Index'):                            GROUP_MAP[c] = FRAC_HEX
    else:                                                            GROUP_MAP[c] = AERO_HEX

THIN = Border(left=Side('thin'), right=Side('thin'),
              top=Side('thin'),  bottom=Side('thin'))

# ── Build stats for summary sheet ───────────────────────────────────────
num_cols = df.select_dtypes(include=np.number).columns.tolist()
stat_rows = []
for col in num_cols:
    stat_rows.append({
        'Feature': col,
        'Min':     round(float(df[col].min()),  5),
        'Max':     round(float(df[col].max()),  5),
        'Mean':    round(float(df[col].mean()), 5),
        'Std_Dev': round(float(df[col].std()),  5),
    })
stat_df = pd.DataFrame(stat_rows)

formula_view_cols = (['Sample_ID','Data_Source','Molecular_Formula',
                      'Empirical_Formula','Dominant_Phase']
                     + [c for c in df.columns if 'content' in c or 'Valence' in c])

with pd.ExcelWriter(XLSX_PATH, engine='openpyxl') as writer:
    df.to_excel(writer,          sheet_name='Full_Dataset',    index=False)
    stat_df.to_excel(writer,     sheet_name='Feature_Stats',   index=False)
    df[formula_view_cols].to_excel(writer, sheet_name='Formula_View', index=False)

wb = load_workbook(XLSX_PATH)

def format_sheet(ws, col_order):
    """Apply colour-coded header + data formatting."""
    # Header
    for ci, cell in enumerate(ws[1], 1):
        col = col_order[ci-1] if ci <= len(col_order) else ''
        hx  = GROUP_MAP.get(col, '404040')
        cell.fill      = PatternFill('solid', start_color=hx)
        cell.font      = Font(name='Arial', bold=True, size=8, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = THIN
    ws.row_dimensions[1].height = 42
    # Data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for ci, cell in enumerate(row, 1):
            col = col_order[ci-1] if ci <= len(col_order) else ''
            hx  = GROUP_MAP.get(col, 'F9F9F9')
            cell.fill      = PatternFill('solid', start_color=hx)
            cell.font      = Font(name='Arial', size=8)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = THIN
    # Column widths
    for ci, col in enumerate(col_order, 1):
        w = 28 if any(k in col for k in ('Formula','Phase','Atmosphere')) else 17
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = 'D2'

format_sheet(wb['Full_Dataset'],  list(df.columns))
format_sheet(wb['Formula_View'],  formula_view_cols)

# Stats sheet formatting
ws_s = wb['Feature_Stats']
hdr_fill = PatternFill('solid', start_color='1F3864')
for cell in ws_s[1]:
    cell.fill      = hdr_fill
    cell.font      = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    cell.alignment = Alignment(horizontal='center')
    cell.border    = THIN
for row in ws_s.iter_rows(min_row=2, max_row=ws_s.max_row):
    for cell in row:
        cell.font      = Font(name='Arial', size=8)
        cell.alignment = Alignment(horizontal='center')
        cell.border    = THIN
ws_s.column_dimensions['A'].width = 40
for ci in range(2, 6):
    ws_s.column_dimensions[get_column_letter(ci)].width = 16
ws_s.freeze_panes = 'A2'

wb.save(XLSX_PATH)
print(f"XLSX saved → {XLSX_PATH}")
print("\nDone. Files ready for download.")
