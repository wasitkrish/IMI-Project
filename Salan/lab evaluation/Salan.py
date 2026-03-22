import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

np.random.seed(42)

# ── Load base file ─────────────────────────────────────────────────────────────
df = pd.read_excel("/home/salan/Documents/SEM_2/python_codes/IMI-Project/krish_sent_files/UHTM_base_200.xlsx", sheet_name="Base Data")

mp  = df["_mp"].values
rho = df["_rho"].values
ve  = df["_ve"].values
en  = df["_en"].values
sp  = df["_sp"].values

def noise(arr, pct):
    return arr * (1 + np.random.normal(0, pct, size=len(arr)))

# ══════════════════════════════════════════════════════════════════════════════
# GROUP E — OXIDATION / CHEMICAL STABILITY
# ══════════════════════════════════════════════════════════════════════════════

# F25 — Oxidation Onset Temperature (K)
# Physics: T_ox = 800 + en*150 + mp*0.05  (stronger/more ionic → resists oxidation longer)
# Relevance: Primary chemical stability criterion; determines passive protection window
F25 = noise(800 + en * 150 + mp * 0.05, 0.05)

# F26 — Parabolic Oxidation Rate Constant kp (kg²/m⁴s)
# Physics: kp = 1e-12 * exp(en*0.8)  — Arrhenius exponential; ionic → less protective oxide
# Relevance: Smaller kp = thinner oxide scale growth = better protection over time
F26 = noise(1e-12 * np.exp(en * 0.8), 0.10)

# F27 — Oxidation Activation Energy (kJ/mol)
# Physics: Ea = 120 + en*30  — higher ionic character raises diffusion barrier
# Relevance: Determines temperature sensitivity of oxidation rate (Arrhenius plot slope)
F27 = noise(120 + en * 30, 0.06)

# F28 — Gravimetric Parabolic Rate (g²/cm⁴s)
# Physics: kp_grav = 1e-10 * en  — proportional to ionic character
# Relevance: Standard gravimetric unit used in thermogravimetric analysis (TGA) experiments
F28 = noise(1e-10 * en, 0.12)

# F29 — Oxide Layer Stability Index (dimensionless)
# Physics: OLS = en*0.7 + ve*0.15  — both bond ionicity and electron density contribute
# Relevance: Composite index for protective oxide scale adherence and self-healing capacity
F29 = noise(en * 0.7 + ve * 0.15, 0.07)

# F30 — Oxygen Diffusivity in Oxide (m²/s)
# Physics: D_O = 1e-14 * exp(-en*1.2)  — lower diffusivity → better oxidation barrier
# Relevance: Controls rate of oxygen transport through protective oxide layer to substrate
F30 = noise(1e-14 * np.exp(-en * 1.2), 0.12)

# ══════════════════════════════════════════════════════════════════════════════
# GROUP F — MICROSTRUCTURAL FEATURES
# ══════════════════════════════════════════════════════════════════════════════

# F31 — Average Grain Size (μm)
# Physics: d = 2 + (sp^-0.3)*10  — higher sintering pressure → finer grains
# Relevance: Hall-Petch effect: smaller grains → higher hardness and fracture resistance
F31 = noise(2 + (sp ** -0.3) * 10, 0.10)

# F32 — Relative Density (%)
# Physics: ρ_rel = min(99.9, 92 + sp*0.18)  — density improves with sintering pressure
# Relevance: High density (>97%) is essential for achieving target strength values
F32 = np.minimum(99.9, noise(92 + sp * 0.18, 0.02))

# F33 — Porosity (%)
# Physics: P = max(0.1, 100 - ρ_rel)  — exact complement of relative density
# Relevance: Pores act as crack initiation sites and oxidation diffusion pathways
F33 = np.maximum(0.1, noise(100 - F32, 0.15))

# F34 — Crystallite Size (nm)
# Physics: D_xrd = grain_size_um * 1000 * 0.08  (XRD coherent domain ~8% of grain size)
# Relevance: Sub-grain domain size from Scherrer analysis; affects dislocation motion
F34 = noise(F31 * 1000 * 0.08, 0.12)

# F35 — Dislocation Density (×10¹² /m²)
# Physics: ρ_disl = 1e12 / d^1.5  — inverse power law with grain size (boundary sources)
# Relevance: Work hardening state and high-temperature creep susceptibility
F35 = noise(1e12 / (F31 ** 1.5), 0.10)

# F36 — Grain Boundary Energy (J/m²)
# Physics: γ_gb = 0.3 + en*0.25 + ve*0.02  (bond character controls boundary strength)
# Relevance: Controls grain boundary sliding at high T; key creep and oxidation pathway
F36 = noise(0.3 + en * 0.25 + ve * 0.02, 0.08)

# ── Build output DataFrame ─────────────────────────────────────────────────────
meta_cols = ["Sample_ID", "Material_System", "Source_Type", "Synthesis_Method", "Crystal_Structure"]
out = df[meta_cols].copy()

out["F25_Oxidation_Onset_Temp_K"]       = F25.round(1)
out["F26_Parabolic_Rate_Const_kg2m4s"]  = F26.round(20)
out["F27_Activation_Energy_Ox_kJmol"]   = F27.round(3)
out["F28_Gravimetric_kp_g2cm4s"]        = F28.round(16)
out["F29_Oxide_Layer_Stability_Idx"]    = F29.round(4)
out["F30_Oxygen_Diffusivity_m2s"]       = F30.round(20)
out["F31_Grain_Size_um"]                = F31.round(3)
out["F32_Relative_Density_pct"]         = F32.round(3)
out["F33_Porosity_pct"]                 = F33.round(4)
out["F34_Crystallite_Size_nm"]          = F34.round(2)
out["F35_Dislocation_Density_1e12_m2"]  = F35.round(4)
out["F36_Grain_Boundary_Energy_Jm2"]    = F36.round(4)

# ── Save to xlsx ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Member3_F25_F36"

NAVY   = "1B2A4A"
RED    = "B91C1C"
GREEN  = "15803D"
WHITE  = "FFFFFF"

OXID_COLS  = ["F25_Oxidation_Onset_Temp_K","F26_Parabolic_Rate_Const_kg2m4s",
              "F27_Activation_Energy_Ox_kJmol","F28_Gravimetric_kp_g2cm4s",
              "F29_Oxide_Layer_Stability_Idx","F30_Oxygen_Diffusivity_m2s"]
MICRO_COLS = ["F31_Grain_Size_um","F32_Relative_Density_pct","F33_Porosity_pct",
              "F34_Crystallite_Size_nm","F35_Dislocation_Density_1e12_m2","F36_Grain_Boundary_Energy_Jm2"]

for ci, col in enumerate(out.columns, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if col in meta_cols:
        cell.fill = PatternFill("solid", start_color=NAVY)
    elif col in OXID_COLS:
        cell.fill = PatternFill("solid", start_color=RED)
    else:
        cell.fill = PatternFill("solid", start_color=GREEN)

EXP = PatternFill("solid", start_color="EFF6FF")
SYN = PatternFill("solid", start_color="FFFBEB")

for ri, row in out.iterrows():
    bg = EXP if row["Source_Type"] == "experimental" else SYN
    for ci, col in enumerate(out.columns, 1):
        cell = ws.cell(row=ri+2, column=ci, value=row[col])
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col in meta_cols:
            cell.fill = bg

for ci, col in enumerate(out.columns, 1):
    ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) * 0.75)
ws.row_dimensions[1].height = 40
ws.freeze_panes = "F2"

wb.save("Salan_member3.xlsx")
print(f"Member 3 done. Shape: {out.shape}")
print(f"Features: {[c for c in out.columns if c.startswith('F')]}")
