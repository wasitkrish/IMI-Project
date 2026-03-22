import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

np.random.seed(42)

# ── Load base file ─────────────────────────────────────────────────────────────
df = pd.read_excel("UHTM_base_200.xlsx", sheet_name="Base Data")

mp  = df["_mp"].values
rho = df["_rho"].values
ve  = df["_ve"].values
en  = df["_en"].values
sp  = df["_sp"].values

def noise(arr, pct):
    return arr * (1 + np.random.normal(0, pct, size=len(arr)))

# ══════════════════════════════════════════════════════════════════════════════
# GROUP C — MECHANICAL PROPERTIES
# ══════════════════════════════════════════════════════════════════════════════

# F13 — Young's Modulus (GPa)
# Physics: E ≈ 200 + mp*0.04 + ve*15  (Gilman-Cohen stiffness-bond-energy relation)
# Relevance: Primary mechanical design parameter for aerodynamic load-bearing structures
F13 = noise(200 + mp * 0.04 + ve * 15, 0.05)

# F14 — Vickers Hardness (GPa)
# Physics: H_v ≈ 15 + en*8 + ve*1.2  (bond character and electron count)
# Relevance: Wear resistance for leading-edge applications; correlates with yield stress
F14 = noise(15 + en * 8 + ve * 1.2, 0.06)

# F15 — Fracture Toughness KIc (MPa√m)
# Physics: KIc ≈ 2.5 + 1.5/en  (ionic bonds = more brittle = lower KIc)
# Relevance: Critical for thermal shock and impact survivability; central to TPS design
F15 = noise(2.5 + 1.5 / en, 0.07)

# F16 — Compressive Strength (GPa)
# Physics: σ_c ≈ 0.6 * E  (empirical ratio for dense ceramics, consistent with data)
# Relevance: Structural components under aerobraking and launch loads
F16 = noise(F13 * 0.6, 0.05)

# F17 — Poisson's Ratio (dimensionless)
# Physics: ν ≈ 0.18 + en*0.02  (covalent materials ~0.18; more ionic → higher ν)
# Relevance: Required for full elastic tensor and FEA simulations
F17 = noise(0.18 + en * 0.02, 0.04)

# F18 — Flexural Strength (MPa)  [Target proxy feature]
# Physics: σ_f = 300 + E*0.8 + H*12 - porosity*15
# porosity estimated as max(0.1, 100 - min(99.9, 92 + sp*0.18))
porosity_est = np.maximum(0.1, 100 - np.minimum(99.9, noise(92 + sp * 0.18, 0.02)))
F18 = noise(300 + F13 * 0.8 + F14 * 12 - porosity_est * 15, 0.05)

# ══════════════════════════════════════════════════════════════════════════════
# GROUP D — THERMAL TRANSPORT FEATURES
# ══════════════════════════════════════════════════════════════════════════════

# F19 — Thermal Conductivity (W/m·K)
# Physics: κ = 20 + ve*4 - en*3  (electronic term ve*4; ionic scattering -en*3)
# Relevance: Determines temperature gradients in re-entry TPS panels
F19 = noise(20 + ve * 4 - en * 3, 0.07)

# F20 — Coefficient of Thermal Expansion (×10⁻⁶/K)
# Physics: CTE = 6.5 + en*0.8 - ve*0.3  (ionicity increases, ve decreases CTE)
# Relevance: Low CTE preferred; mismatch drives thermal stress cracks in composites
F20 = noise(6.5 + en * 0.8 - ve * 0.3, 0.06)

# F21 — Specific Heat Capacity (J/kg·K)
# Physics: Cp ≈ 180 + 800/ρ  (inverse density from Dulong-Petit heavier atoms → lower Cp)
# Relevance: Controls transient thermal response and energy storage during re-entry
F21 = noise(180 + 800 / rho, 0.05)

# F22 — Thermal Diffusivity (m²/s)
# Physics: α = κ / (ρ * Cp)  — exact thermodynamic definition
# Relevance: Governs thermal equilibration speed in TPS panels; directly measurable by LFA
F22 = noise(F19 / (rho * F21 / 1e3), 0.06)

# F23 — Maximum Service Temperature (K)
# Physics: T_max ≈ 0.75 * T_melt  (standard engineering rule for refractory ceramics)
# Relevance: Practical design limit accounting for structural/chemical stability combined
F23 = noise(mp * 0.75, 0.03)

# F24 — Thermal Shock Resistance Parameter (W/m)
# Physics: R = σ_f * κ / (E * CTE)  (Hasselman R-parameter for thermal shock)
# Relevance: Directly predicts resistance to cracking under rapid thermal cycling
F24 = noise((F18 / 1000) * F19 / (F13 * F20 * 1e-6), 0.08)

# ── Build output DataFrame ─────────────────────────────────────────────────────
meta_cols = ["Sample_ID", "Material_System", "Source_Type", "Synthesis_Method", "Crystal_Structure"]
out = df[meta_cols].copy()

out["F13_Youngs_Modulus_GPa"]         = F13.round(2)
out["F14_Vickers_Hardness_GPa"]       = F14.round(3)
out["F15_Fracture_Toughness_MPasm"]   = F15.round(4)
out["F16_Compressive_Strength_GPa"]   = F16.round(3)
out["F17_Poisson_Ratio"]              = F17.round(4)
out["F18_Flexural_Strength_MPa"]      = F18.round(1)
out["F19_Thermal_Conductivity_WmK"]   = F19.round(3)
out["F20_Thermal_Expansion_1e6K"]     = F20.round(4)
out["F21_Specific_Heat_JkgK"]         = F21.round(2)
out["F22_Thermal_Diffusivity_m2s"]    = F22.round(9)
out["F23_Max_Service_Temp_K"]         = F23.round(1)
out["F24_Thermal_Shock_Resistance_Wm"]= F24.round(4)

# ── Save to xlsx ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Member2_F13_F24"

NAVY   = "1B2A4A"
PURPLE = "6D28D9"
ORANGE = "C2410C"
WHITE  = "FFFFFF"

MECH_COLS    = ["F13_Youngs_Modulus_GPa","F14_Vickers_Hardness_GPa","F15_Fracture_Toughness_MPasm",
                "F16_Compressive_Strength_GPa","F17_Poisson_Ratio","F18_Flexural_Strength_MPa"]
THERMAL_COLS = ["F19_Thermal_Conductivity_WmK","F20_Thermal_Expansion_1e6K","F21_Specific_Heat_JkgK",
                "F22_Thermal_Diffusivity_m2s","F23_Max_Service_Temp_K","F24_Thermal_Shock_Resistance_Wm"]

for ci, col in enumerate(out.columns, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if col in meta_cols:
        cell.fill = PatternFill("solid", start_color=NAVY)
    elif col in MECH_COLS:
        cell.fill = PatternFill("solid", start_color=PURPLE)
    else:
        cell.fill = PatternFill("solid", start_color=ORANGE)

EXP = PatternFill("solid", start_color="EFF6FF")
SYN = PatternFill("solid", start_color="FFFBEB")

for ri, row in out.iterrows():
    bg = EXP if row["Source_Type"] == "experimental" else SYN
    for ci, col in enumerate(out.columns, 1):
        cell = ws.cell(row=ri+2, column=ci, value=row[col])
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col in meta_cols:
            cell.fill = bg

for ci, col in enumerate(out.columns, 1):
    ws.column_dimensions[get_column_letter(ci)].width = max(14, len(col) * 0.75)
ws.row_dimensions[1].height = 40
ws.freeze_panes = "F2"

wb.save("Krishh.xlsx")
print(f"Member 2 done. Shape: {out.shape}")
print(f"Features: {[c for c in out.columns if c.startswith('F')]}")
