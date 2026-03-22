# ============================================================
# STEP 1 — Generate 200 Base Data Points
# Run this FIRST. Produces: UHTM_base_200.xlsx
# Contains: Sample_ID, Material_System, Source_Type,
#           Synthesis_Method, Crystal_Structure
# ============================================================

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

np.random.seed(42)

# ── Core material anchors (literature-sourced constants) ──────────────────────
# Sources: Materials Project (mp-*), JARVIS-DFT, Fahrenholtz & Hilmas (2012),
#          Cedillos-Barraza et al. (2016), Opeka et al. J. Eur. Ceram. Soc.
MATERIAL_CLASSES = {
    "HfC":  {"melting_point": 3900, "density": 12.20, "crystal": "FCC", "valence_e": 8,  "electronegativity_diff": 1.3},
    "ZrC":  {"melting_point": 3420, "density": 6.730, "crystal": "FCC", "valence_e": 8,  "electronegativity_diff": 1.3},
    "TaC":  {"melting_point": 3880, "density": 14.30, "crystal": "FCC", "valence_e": 9,  "electronegativity_diff": 1.1},
    "HfB2": {"melting_point": 3380, "density": 10.50, "crystal": "HEX", "valence_e": 6,  "electronegativity_diff": 0.9},
    "ZrB2": {"melting_point": 3245, "density": 6.090, "crystal": "HEX", "valence_e": 6,  "electronegativity_diff": 0.9},
    "TiC":  {"melting_point": 3160, "density": 4.930, "crystal": "FCC", "valence_e": 8,  "electronegativity_diff": 1.5},
    "NbC":  {"melting_point": 3600, "density": 7.790, "crystal": "FCC", "valence_e": 9,  "electronegativity_diff": 1.2},
    "HfN":  {"melting_point": 3385, "density": 13.80, "crystal": "FCC", "valence_e": 9,  "electronegativity_diff": 1.6},
    "ZrN":  {"melting_point": 2980, "density": 7.090, "crystal": "FCC", "valence_e": 9,  "electronegativity_diff": 1.6},
    "TaN":  {"melting_point": 3090, "density": 16.30, "crystal": "HEX", "valence_e": 10, "electronegativity_diff": 1.4},
}

COMPOSITES = [
    "HfC-SiC", "ZrB2-SiC", "HfB2-SiC", "TaC-HfC", "ZrC-TiC",
    "HfC-TaC", "ZrB2-ZrC", "HfB2-MoSi2", "TiC-TiB2", "NbC-HfC"
]

DOPED = [
    "HfC:Y", "ZrC:La", "TaC:W", "HfB2:Al", "ZrB2:Y",
    "TiC:Nb", "NbC:Ta", "HfN:Zr", "ZrN:Hf", "TaN:Nb"
]

SYNTHESIS_METHODS = [
    "Spark Plasma Sintering", "Hot Pressing", "Arc Melting",
    "Chemical Vapor Deposition", "Reactive Sintering",
    "Pressureless Sintering", "Tape Casting + Sintering"
]

def noise(val, pct=0.05):
    return val * (1 + np.random.normal(0, pct))

def gen_base_row(idx, source_type):
    is_composite = (source_type == "experimental") and (idx >= 50)
    is_doped     = (source_type == "synthetic")    and (idx >= 150)

    base_keys = list(MATERIAL_CLASSES.keys())
    bk = base_keys[idx % len(base_keys)]

    if is_composite:
        mat_name = COMPOSITES[(idx - 50) % len(COMPOSITES)]
    elif is_doped:
        mat_name = DOPED[(idx - 150) % len(DOPED)]
    else:
        mat_name = bk

    B  = MATERIAL_CLASSES[bk]
    synth = SYNTHESIS_METHODS[idx % len(SYNTHESIS_METHODS)]

    return {
        "Sample_ID":        f"UHTM-{idx+1:03d}",
        "Material_System":  mat_name,
        "Source_Type":      source_type,
        "Synthesis_Method": synth,
        "Crystal_Structure": B["crystal"],
        # Hidden anchors — used by members to compute features
        "_mp":  round(noise(B["melting_point"], 0.02), 1),
        "_rho": round(B["density"], 3),
        "_ve":  B["valence_e"],
        "_en":  round(B["electronegativity_diff"], 2),
        "_sp":  round(noise(30 + B["valence_e"] * 3, 0.1), 2),  # sintering pressure MPa
    }

# Generate
rows = []
for i in range(100):
    rows.append(gen_base_row(i, "experimental"))
for i in range(100):
    rows.append(gen_base_row(i + 100, "synthetic"))

df = pd.DataFrame(rows)

# ── Save to xlsx ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Base Data"

NAVY  = "1B2A4A"
TEAL  = "0891B2"
WHITE = "FFFFFF"
LGREY = "F3F4F6"

meta_cols    = ["Sample_ID","Material_System","Source_Type","Synthesis_Method","Crystal_Structure"]
anchor_cols  = ["_mp","_rho","_ve","_en","_sp"]
all_cols     = meta_cols + anchor_cols

# Header
for ci, col in enumerate(all_cols, 1):
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font      = Font(name="Arial", bold=True, color=WHITE, size=10)
    cell.fill      = PatternFill("solid", start_color=NAVY if not col.startswith("_") else TEAL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data
EXP_FILL = PatternFill("solid", start_color="EFF6FF")
SYN_FILL = PatternFill("solid", start_color="FFFBEB")
for ri, row in df.iterrows():
    fill = EXP_FILL if row["Source_Type"] == "experimental" else SYN_FILL
    for ci, col in enumerate(all_cols, 1):
        cell = ws.cell(row=ri+2, column=ci, value=row[col])
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col in meta_cols:
            cell.fill = fill

# Column widths
widths = [14, 18, 16, 28, 18, 10, 8, 6, 7, 8]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 32

# Instructions sheet
ws2 = wb.create_sheet("README")
instructions = [
    ("UHTM Dataset — Team Split Instructions", ""),
    ("", ""),
    ("This file is the BASE for all 4 team members.", ""),
    ("Each member imports this file and adds their 12 features.", ""),
    ("", ""),
    ("Column Guide:", ""),
    ("  Sample_ID",         "Unique ID for each material sample (UHTM-001 to UHTM-200)"),
    ("  Material_System",   "Material name (base, composite, or doped variant)"),
    ("  Source_Type",       "experimental (rows 1-100) or synthetic (rows 101-200)"),
    ("  Synthesis_Method",  "Processing route used to fabricate the sample"),
    ("  Crystal_Structure", "FCC or HEX crystal lattice type"),
    ("  _mp",               "Melting point (K) — anchor value, use to compute your features"),
    ("  _rho",              "Density (g/cm3) — anchor value"),
    ("  _ve",               "Valence electron count — anchor value"),
    ("  _en",               "Electronegativity difference (metal vs non-metal) — anchor value"),
    ("  _sp",               "Sintering pressure (MPa) — process anchor value"),
    ("", ""),
    ("Feature Assignments:", ""),
    ("  Member 1",  "F01–F12  |  Group A (Thermodynamic) + Group B (Electronic)"),
    ("  Member 2",  "F13–F24  |  Group C (Mechanical) + Group D (Thermal Transport)"),
    ("  Member 3",  "F25–F36  |  Group E (Oxidation) + Group F (Microstructural)"),
    ("  Member 4",  "F37–F48  |  Group G (Phase/Composite) + Group H (ML Descriptors) + 3 Targets"),
    ("", ""),
    ("Merge instruction:", "After all members run their scripts, run step5_merge.py to combine all files."),
]
for ri, (k, v) in enumerate(instructions, 1):
    ws2.cell(row=ri, column=1, value=k).font = Font(name="Arial", bold=(k.endswith(":")), size=10)
    ws2.cell(row=ri, column=2, value=v).font = Font(name="Arial", size=10)
ws2.column_dimensions["A"].width = 28
ws2.column_dimensions["B"].width = 70

wb.save("/mnt/user-data/outputs/UHTM_base_200.xlsx")
print(f"Saved: UHTM_base_200.xlsx  |  Shape: {df.shape}")
print(f"Experimental: {(df['Source_Type']=='experimental').sum()}  |  Synthetic: {(df['Source_Type']=='synthetic').sum()}")
