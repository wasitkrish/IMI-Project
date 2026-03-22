import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

META = ["Sample_ID", "Material_System", "Source_Type", "Synthesis_Method", "Crystal_Structure"]

# ── Load all member files ──────────────────────────────────────────────────────
d1 = pd.read_excel("AadiDev.xlsx")
d2 = pd.read_excel("Krishh.xlsx")
d3 = pd.read_excel("Salan.xlsx")
d4 = pd.read_excel("Niranjan.xlsx")

# ── Validate all members have 200 rows ────────────────────────────────────────
for name, df in [("Member1", d1), ("Member2", d2), ("Member3", d3), ("Member4", d4)]:
    assert len(df) == 200, f"{name} has {len(df)} rows, expected 200"
    assert list(df["Sample_ID"]) == list(d1["Sample_ID"]), f"{name} Sample_ID mismatch"
    print(f"{name}: {len(df)} rows, {len([c for c in df.columns if c.startswith('F')])} features OK")

# ── Merge ──────────────────────────────────────────────────────────────────────
f1_cols = [c for c in d1.columns if c.startswith("F")]
f2_cols = [c for c in d2.columns if c.startswith("F")]
f3_cols = [c for c in d3.columns if c.startswith("F")]
f4_cols = [c for c in d4.columns if c.startswith("F")]
t_cols  = [c for c in d4.columns if c.startswith("T")]

final = d1[META + f1_cols].copy()
final = final.join(d2[f2_cols])
final = final.join(d3[f3_cols])
final = final.join(d4[f4_cols + t_cols])

print(f"\nFinal merged shape: {final.shape}")
print(f"Features: {len([c for c in final.columns if c.startswith('F')])} | Targets: {len(t_cols)}")

# ── Save final xlsx with colour coding by group ───────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "UHTM_Full_Dataset"

# Colour map per feature group
def col_color(col):
    if col in META:             return ("1B2A4A", "FFFFFF")   # navy  / white
    n = int(col[1:3]) if col.startswith("F") and col[1:3].isdigit() else 0
    if  1 <= n <= 6:  return ("0891B2", "FFFFFF")  # teal   — Group A Thermodynamic
    if  7 <= n <= 12: return ("0284C7", "FFFFFF")  # blue   — Group B Electronic
    if 13 <= n <= 18: return ("6D28D9", "FFFFFF")  # purple — Group C Mechanical
    if 19 <= n <= 24: return ("C2410C", "FFFFFF")  # orange — Group D Thermal
    if 25 <= n <= 30: return ("B91C1C", "FFFFFF")  # red    — Group E Oxidation
    if 31 <= n <= 36: return ("15803D", "FFFFFF")  # green  — Group F Microstructural
    if 37 <= n <= 42: return ("3730A3", "FFFFFF")  # indigo — Group G Phase
    if 43 <= n <= 48: return ("0F766E", "FFFFFF")  # teal   — Group H ML
    if col.startswith("T"):  return ("7F1D1D", "FFFFFF")  # dark red — Targets
    return ("374151", "FFFFFF")

for ci, col in enumerate(final.columns, 1):
    bg, fg = col_color(col)
    cell = ws.cell(row=1, column=ci, value=col)
    cell.font      = Font(name="Arial", bold=True, color=fg, size=9)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

EXP = PatternFill("solid", start_color="EFF6FF")
SYN = PatternFill("solid", start_color="FFFBEB")

for ri, row in final.iterrows():
    bg_row = EXP if row["Source_Type"] == "experimental" else SYN
    for ci, col in enumerate(final.columns, 1):
        cell = ws.cell(row=ri+2, column=ci, value=row[col])
        cell.font      = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if col in META:
            cell.fill = bg_row

for ci, col in enumerate(final.columns, 1):
    ws.column_dimensions[get_column_letter(ci)].width = max(13, len(col) * 0.72)
ws.row_dimensions[1].height = 44
ws.freeze_panes = "F2"

# ── Summary stats sheet ───────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary Stats")
num_cols = [c for c in final.columns if c.startswith("F") or c.startswith("T")]
stats = final[num_cols].apply(pd.to_numeric, errors="coerce").describe().T.reset_index()
stats.columns = ["Feature"] + list(stats.columns[1:])
for ci, h in enumerate(stats.columns, 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    cell.fill = PatternFill("solid", start_color="1B2A4A")
    cell.alignment = Alignment(horizontal="center")
for ri, row in stats.iterrows():
    for ci, val in enumerate(row, 1):
        cell = ws2.cell(row=ri+2, column=ci, value=round(val, 6) if isinstance(val, float) else val)
        cell.font = Font(name="Arial", size=9)
        cell.fill = PatternFill("solid", start_color="F3F4F6" if ri % 2 == 0 else "FFFFFF")
ws2.column_dimensions["A"].width = 36
for ci in range(2, len(stats.columns)+1):
    ws2.column_dimensions[get_column_letter(ci)].width = 14

# ── Feature legend sheet ──────────────────────────────────────────────────────
ws3 = wb.create_sheet("Feature Legend")
legend = [
    ("Group",           "Features", "Colour",     "Member", "Physical Domain"),
    ("A: Thermodynamic","F01–F06",  "Teal",       "M1",     "Phase stability, bonding"),
    ("B: Electronic",   "F07–F12",  "Blue",       "M1",     "Quantum/DFT properties"),
    ("C: Mechanical",   "F13–F18",  "Purple",     "M2",     "Structural integrity"),
    ("D: Thermal",      "F19–F24",  "Orange",     "M2",     "Heat transport"),
    ("E: Oxidation",    "F25–F30",  "Red",        "M3",     "Chemical stability"),
    ("F: Microstructural","F31–F36","Green",      "M3",     "Process–structure link"),
    ("G: Phase/Composite","F37–F42","Indigo",     "M4",     "Multi-phase systems"),
    ("H: ML Descriptors","F43–F48", "Teal-dark",  "M4",     "Pareto/reward signals"),
    ("Targets",          "T1–T3",   "Dark Red",   "M4",     "Flexural str, Oxidation, Thermal shock"),
]
HDR_COLORS = ["1B2A4A","0891B2","0284C7","6D28D9","C2410C","B91C1C","15803D","3730A3","0F766E","7F1D1D"]
for ri, row in enumerate(legend, 1):
    for ci, val in enumerate(row, 1):
        cell = ws3.cell(row=ri, column=ci, value=val)
        if ri == 1:
            cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", start_color="1B2A4A")
        else:
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color="F3F4F6" if ri % 2 == 0 else "FFFFFF")
        cell.alignment = Alignment(horizontal="left", vertical="center")
for ci, w in enumerate([28, 16, 14, 10, 36], 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

wb.save("UHTM_final_200x48.xlsx")
print("\nSaved: UHTM_final_200x48.xlsx")
print(f"Total columns: {len(final.columns)}  (5 meta + 48 features + 3 targets)")
